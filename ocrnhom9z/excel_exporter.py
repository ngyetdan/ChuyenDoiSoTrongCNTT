# excel_exporter.py - Xuất dữ liệu ra Excel

import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import *

class ExcelExporter:
    """Class xử lý xuất dữ liệu ra Excel"""
    
    def __init__(self):
        pass
    
    def export_to_excel(self, df, file_path):
        """Xuất DataFrame ra file Excel với định dạng đẹp"""
        if df.empty:
            raise ValueError("Không có dữ liệu để xuất")
        
        try:
            # Tạo tên sheet với timestamp
            sheet_name = f"{EXCEL_SHEET_NAME_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Loại bỏ các cột không có dữ liệu
            df_export = self._remove_empty_columns(df)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Xuất dữ liệu
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)

                # Lấy worksheet để định dạng
                worksheet = writer.sheets[sheet_name]

                # Áp dụng định dạng
                self._format_worksheet(worksheet, df_export)
                
            return True, f"Đã xuất thành công {len(df)} sinh viên"
            
        except Exception as e:
            return False, f"Lỗi xuất Excel: {str(e)}"

    def _remove_empty_columns(self, df):
        """Loại bỏ các cột không có dữ liệu"""
        df_clean = df.copy()

        # Danh sách các cột có thể bị trống
        potential_empty_cols = ['KT2', 'KDT', 'KT3', 'KT4', 'TH1', 'TH2', 'TH3', 'TH4']

        columns_to_remove = []

        for col in potential_empty_cols:
            if col in df_clean.columns:
                # Convert sang string và kiểm tra
                col_str = df_clean[col].fillna('').astype(str)

                # Kiểm tra các giá trị có thể coi là "trống"
                empty_indicators = ['', 'nan', 'None', '0.0', '1.0', 'NaN']

                # Đếm số giá trị thực sự có nghĩa
                meaningful_values = []
                for val in col_str:
                    val_clean = str(val).strip()
                    if val_clean not in empty_indicators:
                        meaningful_values.append(val_clean)

                # Nếu không có giá trị có nghĩa nào, loại bỏ cột
                if len(meaningful_values) == 0:
                    columns_to_remove.append(col)

        # Loại bỏ các cột trống
        if columns_to_remove:
            df_clean = df_clean.drop(columns=columns_to_remove)
            print(f"🗑️ Đã loại bỏ {len(columns_to_remove)} cột không có dữ liệu: {columns_to_remove}")

        return df_clean

    def export_raw_data(self, df, file_path):
        """Xuất dữ liệu nguyên bản không qua validation"""
        if df.empty:
            raise ValueError("Không có dữ liệu để xuất")

        try:
            # Tạo tên sheet với timestamp
            sheet_name = f"Raw_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # Loại bỏ các cột không có dữ liệu
            df_export = self._remove_empty_columns(df)

            # Chỉ đảm bảo NaN được hiển thị như ô trống
            df_export = df_export.fillna('')

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Xuất dữ liệu nguyên bản
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)

                # Lấy worksheet để định dạng cơ bản
                worksheet = writer.sheets[sheet_name]

                # Chỉ áp dụng định dạng cơ bản (header và border)
                self._format_raw_worksheet(worksheet, df_export)

            return True, f"Đã xuất dữ liệu nguyên bản {len(df)} sinh viên"

        except Exception as e:
            return False, f"Lỗi xuất dữ liệu nguyên bản: {str(e)}"

    def _format_raw_worksheet(self, worksheet, df):
        """Định dạng cơ bản cho worksheet dữ liệu nguyên bản"""
        # Chỉ định dạng header
        self._format_header(worksheet)

        # Tự động điều chỉnh độ rộng cột
        self._auto_adjust_columns(worksheet)

        # Thêm border cơ bản
        self._add_borders(worksheet, df)
    
    def _format_worksheet(self, worksheet, df):
        """Định dạng worksheet Excel"""
        # Định dạng header
        self._format_header(worksheet)
        
        # Tự động điều chỉnh độ rộng cột
        self._auto_adjust_columns(worksheet)
        
        # Định dạng dữ liệu
        self._format_data_cells(worksheet, df)
        
        # Thêm border
        self._add_borders(worksheet, df)
    
    def _format_header(self, worksheet):
        """Định dạng header"""
        header_font = Font(
            name='Arial',
            size=12,
            bold=True,
            color='FFFFFF'
        )
        
        header_fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # Áp dụng định dạng cho hàng đầu tiên
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_adjust_columns(self, worksheet):
        """Tự động điều chỉnh độ rộng cột"""
        column_widths = {
            'A': 8,   # STT
            'B': 15,  # Lớp
            'C': 15,  # MSV
            'D': 20,  # Họ và đệm
            'E': 12,  # Tên
            'F': 8,   # CC
            'G': 8    # KT1
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
    
    def _format_data_cells(self, worksheet, df):
        """Định dạng các ô dữ liệu"""
        data_font = Font(name='Arial', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Định dạng từng cột
        for row_idx in range(2, len(df) + 2):  # Bắt đầu từ hàng 2
            # STT - center
            worksheet[f'A{row_idx}'].font = data_font
            worksheet[f'A{row_idx}'].alignment = center_alignment
            
            # Lớp - center
            worksheet[f'B{row_idx}'].font = data_font
            worksheet[f'B{row_idx}'].alignment = center_alignment
            
            # MSV - center
            worksheet[f'C{row_idx}'].font = data_font
            worksheet[f'C{row_idx}'].alignment = center_alignment
            
            # Họ và đệm - left
            worksheet[f'D{row_idx}'].font = data_font
            worksheet[f'D{row_idx}'].alignment = left_alignment
            
            # Tên - left
            worksheet[f'E{row_idx}'].font = data_font
            worksheet[f'E{row_idx}'].alignment = left_alignment
            
            # Điểm CC - center
            worksheet[f'F{row_idx}'].font = data_font
            worksheet[f'F{row_idx}'].alignment = center_alignment
            
            # Điểm KT1 - center
            worksheet[f'G{row_idx}'].font = data_font
            worksheet[f'G{row_idx}'].alignment = center_alignment
    
    def _add_borders(self, worksheet, df):
        """Thêm border cho bảng"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Áp dụng border cho tất cả ô có dữ liệu
        for row_idx in range(1, len(df) + 2):
            for col_idx in range(1, len(EXCEL_HEADERS) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
    
    def create_summary_sheet(self, writer, df, sheet_name):
        """Tạo sheet tóm tắt thống kê"""
        try:
            summary_data = self._generate_summary_data(df)
            
            # Tạo DataFrame cho summary
            summary_df = pd.DataFrame(summary_data)
            
            # Xuất summary sheet
            summary_sheet_name = f"ThongKe_{sheet_name}"
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
            
            return True
        except Exception as e:
            print(f"Lỗi tạo summary sheet: {e}")
            return False
    
    def _generate_summary_data(self, df):
        """Tạo dữ liệu thống kê"""
        summary = []
        
        # Thống kê tổng quan
        summary.append(["Thống kê", "Giá trị"])
        summary.append(["Tổng số sinh viên", len(df)])
        summary.append(["", ""])
        
        # Thống kê theo lớp
        if 'Lớp' in df.columns:
            class_counts = df['Lớp'].value_counts()
            summary.append(["Phân bố theo lớp", ""])
            for class_name, count in class_counts.items():
                summary.append([class_name, count])
            summary.append(["", ""])
        
        # Thống kê điểm
        try:
            if 'CC' in df.columns:
                cc_scores = pd.to_numeric(df['CC'], errors='coerce').dropna()
                if len(cc_scores) > 0:
                    summary.append(["Thống kê điểm CC", ""])
                    summary.append(["Trung bình", f"{cc_scores.mean():.2f}"])
                    summary.append(["Cao nhất", f"{cc_scores.max():.1f}"])
                    summary.append(["Thấp nhất", f"{cc_scores.min():.1f}"])
                    summary.append(["Điểm 10", (cc_scores == 10).sum()])
                    summary.append(["Điểm 0", (cc_scores == 0).sum()])
                    summary.append(["", ""])
            
            if 'KT1' in df.columns:
                kt1_scores = pd.to_numeric(df['KT1'], errors='coerce').dropna()
                if len(kt1_scores) > 0:
                    summary.append(["Thống kê điểm KT1", ""])
                    summary.append(["Trung bình", f"{kt1_scores.mean():.2f}"])
                    summary.append(["Cao nhất", f"{kt1_scores.max():.1f}"])
                    summary.append(["Thấp nhất", f"{kt1_scores.min():.1f}"])
                    summary.append(["Điểm 10", (kt1_scores == 10).sum()])
                    summary.append(["Điểm 0", (kt1_scores == 0).sum()])
        except Exception as e:
            print(f"Lỗi tính thống kê điểm: {e}")
        
        return summary
